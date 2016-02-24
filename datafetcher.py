
#Once you’ve imported the openpyxl module, you’ll be able to use the openpyxl.load_workbook()function and other such functions from openpyxl library

#Opening Excel Documents with OpenPyXL
import openpyxl
wb = openpyxl.load_workbook('example.xlsx') #This Workbook object represents the Excel file, a bit like how a File object represents an opened text file.
type(wb)

#Getting Sheets from the Workbook
import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet3')
sheet
type(sheet) <class 'openpyxl.worksheet.worksheet.Worksheet'>
sheet.title
anotherSheet = wb.get_active_sheet()
anotherSheet

#Getting Cells from the Sheets
import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
sheet['A1']
sheet['A1'].value
c = sheet['B1']
c.value
'Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value
'Cell ' + c.coordinate + ' is ' + c.value
sheet['C1'].value

#Converting Between Column Letters and Numbers
import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
get_column_letter(1)
get_column_letter(2)
get_column_letter(27)
get_column_letter(900)
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
get_column_letter(sheet.get_highest_column())
column_index_from_string('A')
column_index_from_string('AA')

#Getting Rows and Columns from the Sheets
 import openpyxl
 wb = openpyxl.load_workbook('example.xlsx')
 sheet = wb.get_sheet_by_name('Sheet1')
 tuple(sheet['A1':'C3'])
for rowOfCellObjects in sheet['A1':'C3']:
 for cellObj in rowOfCellObjects:
   print(cellObj.coordinate, cellObj.value)
	print('--- END OF ROW ---')

#Write the Results to a File
# Open a new text file (fetchdata.tx) and write the contents of example.xlsx to it.
print('Writing results...')
resultFile = open('fetchdata.txt', 'w')
resultFile.write('allData = ' + pprint.pformat(example.xlsx))
resultFile.close()
print('Done.')	
	
	
	
	
	
	
	
	
	
	
	
	
